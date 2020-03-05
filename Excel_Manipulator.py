import os
import shutil
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font


# Function used to add "Case ID" to first column of each Excel file, and to create a master file recording "Case ID"
# and total number of compressions
def run_excel_manipulation(path):
    file_list = os.listdir(path)
    # If the master .xlsx file is open, abort the script and ask the user to close it to avoid an exception
    if "Compression_Master_Data_File.xlsx" in file_list:
        try:
            os.rename(path + "\\" + "Compression_Master_Data_File.xlsx", path + "\\" +
                      "Compression_Master_Data_File.xlsx")
        except PermissionError:
            print("\nMaster Data Excel file is open.  Aborting.  Please close and re-run script.")
            return

    # Ensure only .xlsx, .xls, .xlsb, and .xlsm are included
    filtered_list = []
    others_list = []
    master_file_flag = False
    for i in range(len(file_list)):
        list_element = file_list[i]
        extension = list_element[len(list_element) - 5: len(list_element)]
        if list_element == "Compression_Master_Data_File.xlsx":
            master_file_flag = True
            continue
        elif list_element[len(list_element) - 5] != '.' or list_element[0] == '~':
            continue
        elif extension == ".xls" or extension == ".xlsx" or extension == ".xlsb" or extension == ".xlsm":
            filtered_list.append(list_element)
        else:
            others_list.append(list_element)

    # If there are no .xlsx, .xls, .xlsb, or .xlsm files in the folder:
    if len(filtered_list) == 0:
        print("\nNo Excel files to manipulate.  Aborting Script.")
        return

    # Create directories to place all processed Excel files as well as all non-relevant files
    excel_directory_name = create_directory(path, "Processed_Excel_Files", "Excel")
    others_directory_name = create_directory(path, "Non_Excel_Files", "Non-Excel")

    # Move all non_Excel files to specified directory
    for other_file in others_list:
        try:
            shutil.move(path + "\\" + other_file, others_directory_name + "\\" + other_file)
        except PermissionError:
            print("\nFile %s is open.  Skipping over file.  Please re-run script once file is closed." % other_file)

    # Create master data file if it hasn't yet been created, or open it if it has
    if not master_file_flag:
        master_workbook = openpyxl.Workbook()
        master_sheet = master_workbook.active
        master_sheet.cell(row=1, column=1).value = "Case ID"
        master_sheet.cell(row=1, column=2).value = "Compression Count"
        for p in range(1, 3):
            master_sheet.cell(row=1, column=p).font = Font(bold=True)
            master_sheet.column_dimensions[get_column_letter(p)].width = 38
    else:
        master_workbook = openpyxl.load_workbook(filename=path + "\\" + "Compression_Master_Data_File.xlsx")
        master_sheet = master_workbook.active

    # Open each Excel file and add the Case ID to column A, then move the file to the specified directory
    print("\nOne moment please...")
    for file in filtered_list:
        file_path = path + "\\" + file
        try:
            os.rename(file_path, file_path)

            # Edit the file name to only include what's needed and store in a variable
            case_id = remove_extra_characters(file)

            # Open Excel File
            workbook = openpyxl.load_workbook(filename=file_path)
            worksheet = workbook.active
            compression_count = worksheet.max_row - 1

            # Check 'Case ID' against the values in the master data sheet to see this file has already been processed
            is_repeat = False
            for n in range(2, master_sheet.max_row + 1):
                if case_id == master_sheet.cell(row=n, column=1).value:
                    # If 'Case ID' matches, confirm a repeat using 'Compression Count'
                    if compression_count == master_sheet.cell(row=n, column=2).value:
                        is_repeat = True
            if is_repeat:
                print("\nFile %s has already been processed.  Adding it back into the specified directory." % file)
                workbook.close()
                shutil.move(file_path, excel_directory_name + "\\" + file)
                continue

            # Create new column A and add 'Case ID' for every row with data
            worksheet.insert_cols(1)
            worksheet.cell(row=1, column=1).value = "Case ID"
            for j in range(2, compression_count + 2):
                worksheet.cell(row=j, column=1).value = case_id

            # Make all titles bold and standardize column dimensions
            for k in range(1, 9):
                worksheet.cell(row=1, column=k).font = Font(bold=True)
                worksheet.column_dimensions[get_column_letter(k)].width = 20
            workbook.save(file_path)
            workbook.close()

            # Move Excel file to appropriate folder
            shutil.move(file_path, excel_directory_name + "\\" + file)

            # Add 'Case ID' and 'Compression Count' to master data file
            final_row = master_sheet.max_row + 1
            master_sheet.cell(row=final_row, column=1).value = case_id
            master_sheet.cell(row=final_row, column=2).value = compression_count
        except PermissionError:
            print("\nFile %s is open.  Skipping over file.  Please re-run script once file is closed." % file)

    # Save Master Data File
    master_workbook.save(filename=path + "\\" + "Compression_Master_Data_File.xlsx")
    master_workbook.close()
    return


# Create a directory in the specified path if one hasn't already been created
def create_directory(path, file_name, file_type):
    directory_name = path + "\\" + file_name
    try:
        os.mkdir(directory_name)
        print("\nSuccessfully created the directory %s to store %s files." % (directory_name, file_type))
    except OSError:
        print("\n%s files are stored in the directory %s." % (file_type, directory_name))
    finally:
        return directory_name


# Remove un-needed characters from the end of the file name
def remove_extra_characters(file_name):
    reversed_file_name = ""
    # Reverse string to find last instance of '_'
    for k in range(len(file_name), 0, -1):
        reversed_file_name += file_name[k - 1]
    underscore_marker = reversed_file_name.find("_")
    # If no '_', just eliminate file extension
    if underscore_marker == -1:
        chars_to_remove = file_name[len(file_name) - 5: len(file_name)]
        return file_name.replace(chars_to_remove, "")
    # If '_' has a character to its left, it doesn't have extra digits
    else:
        try:
            char_or_int = int(file_name[len(file_name) - 1 - underscore_marker - 1])
            has_extra_digits = True
        except ValueError:
            has_extra_digits = False
        # If there are extra digits, remove them along with the file extension
        if has_extra_digits:
            chars_to_remove = file_name[len(file_name) - 1 - underscore_marker: len(file_name)]
            return file_name.replace(chars_to_remove, "")
        else:
            chars_to_remove = file_name[len(file_name) - 5: len(file_name)]
            return file_name.replace(chars_to_remove, "")


run_excel_manipulation(r"C:\Users\mnarcisi\Documents\Mike\Scientific Affairs\Excel_File_Testing\Main AY1")
