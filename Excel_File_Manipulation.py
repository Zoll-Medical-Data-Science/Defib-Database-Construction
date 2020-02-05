import os
import openpyxl
from openpyxl.utils import get_column_letter


#  This function is used to add a column to excel based data with the title of the sheet as the volue for every row in the column
def add_title_column(path):
    file_list = os.listdir(path)

    # Ensure only .xlsx, .xls, .xlsb, and .xlsm files are shown to user
    user_list = []
    for i in range(len(file_list)):
        list_element = file_list[i]
        extension = list_element[len(list_element) - 5: len(list_element)]
        if (extension == ".xls" or extension == ".xlsx" or extension == ".xlsb" or extension == ".xlsm") and list_element.find("(Column_Added)") == -1:
            user_list.append(list_element)

    # Show user the relevant files
    user_choice = 0
    while user_choice < 1 or user_choice > len(user_list):
        print('\033[1m' + "\nPlease choose which Excel file to manipulate (Enter number only):\n")
        for j in range(len(user_list)):
            print("(" + str(j + 1) + ") " + user_list[j])
        user_choice = input()
        try:
            user_choice = int(user_choice)
            if user_choice < 1 or user_choice > len(user_list):
                print("\nPlease select one of the given selections")
        except ValueError:
            print("\nPlease enter a numerical value")
            user_choice = 0
    selected_file_name = path + "\\" + user_list[user_choice - 1]

    # Remove '.xlsx' or other extension and trailing digits from the end of the file name
    formatted_file_name = user_list[user_choice - 1]
    updated_name_format = ""
    reversed_file_name = ""
    underscore_marker = -1
    is_extra_digit = False
    # Reverse string to find last instance of '_'
    for k in range(len(formatted_file_name), 0, -1):
        reversed_file_name = reversed_file_name + formatted_file_name[k - 1]
    underscore_marker = reversed_file_name.find("_")
    # If '_' has a character to its left, it doesn't have extra digits
    if underscore_marker != -1:
        try:
            character_or_integer = int(formatted_file_name[len(formatted_file_name) - 1 - underscore_marker - 1])
            is_extra_digit = True
        except ValueError:
            is_extra_digit = False
        # If there are extra digits, remove them along with the file extension
        if is_extra_digit:
            text_to_remove = formatted_file_name[
                             len(formatted_file_name) - 1 - underscore_marker: len(formatted_file_name)]
            updated_name_format = formatted_file_name.replace(text_to_remove, "")
        else:
            text_to_remove = formatted_file_name[
                             len(formatted_file_name) - 5: len(formatted_file_name)]
            updated_name_format = formatted_file_name.replace(text_to_remove, "")
    else:
        text_to_remove = formatted_file_name[
                         len(formatted_file_name) - 5: len(formatted_file_name)]
        updated_name_format = formatted_file_name.replace(text_to_remove, "")

    # Open the selected Excel file and add an additional column with the workbook name in it for all rows
    excel = openpyxl.load_workbook(path + "\\" + formatted_file_name)
    sheet = excel.active
    high_column = sheet.max_column
    high_row = sheet.max_row
    # Open a new workbook
    new_Excel = openpyxl.Workbook()
    new_sheet = new_Excel.active
    # Copy all columns up to the blank column into the new workbook, and add title to each row
    for column in range(1, high_column + 2):
        for row in range(1, high_row + 1):
            if column == high_column + 1:
                if row == 1:
                    new_sheet.cell(row=row, column=column).value = "File Name"
                else:
                    new_sheet.cell(row=row, column=column).value = updated_name_format
            else:
                new_sheet.cell(row = row, column = column).value = sheet.cell(row = row, column = column).value
        column_letter = get_column_letter(column)
        new_sheet.column_dimensions[column_letter].width = 17
    # Save new workbook
    new_Excel.save(path + "\\" + formatted_file_name[0 : len(formatted_file_name) - 5] + "_(Column_Added)" +
                   formatted_file_name[len(formatted_file_name) - 5: len(formatted_file_name)])

    print("\nPlease check the source folder for your new file.")
