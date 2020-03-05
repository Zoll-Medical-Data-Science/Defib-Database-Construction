import os
import shutil
import pandas as pd
import csv
import openpyxl
from openpyxl.styles import Font
import statistics

#  This function is used to filter all text files in a folder of sample defib data and pick out all "defib shock"
#  rows and their times
def run_data_manipulation(path):
    print("\n\033[1m" + "Current Directory: " + path)
    file_list = os.listdir(path)

    # If the master .xlsx file is open, abort the script and ask the user to close it to avoid an exception
    if "Defib_Shock_Master_Data_File.xlsx" in file_list:
        try:
            os.rename(path + "\\" + "Defib_Shock_Master_Data_File.xlsx", path + "\\" +
                      "Defib_Shock_Master_Data_File.xlsx")
        except PermissionError:
            print("\nMaster Data Excel file is open.  Aborting.  Please close and re-run script.")
            return "N/A", "N/A"

    # Ensure only .txt files are included
    user_list = []
    for i in range(len(file_list)):
        list_element = file_list[i]
        if (list_element[len(list_element) - 4: len(list_element)] == ".txt" or list_element[len(list_element) - 4:
        len(list_element)] == ".log") and (list_element != "Defib_Shock_Extracted_Data.txt" and list_element !=
                                           "Defib_Shock_Case_Tracker.txt"):
            user_list.append(list_element)

    # If there are no .txt or .log files in the folder:
    if len(user_list) == 0:
        # If the output files have already been created
        if ("Defib_Shock_Master_Data_File.xlsx" in file_list) or ("Defib_Shock_Extracted_Data.txt" in file_list) or \
                ("Defib_Shock_Case_Tracker.txt" in file_list):
            print("\nNo .txt or .log files to manipulate.  Updating output files and ending the script.")
            return path + "\\" + "Defib_Shock_Extracted_Data.txt", path + "\\" + "Defib_Shock_Case_Tracker.txt"

        # If the script hasn't successfully run in the folder yet
        else:
            print("\nNo .txt or .log files to manipulate.  Aborting Script.")
            return "N/A", "N/A"

    # Create directories to place all processed .txt and .log files (if they haven't already been created)
    text_directory_name = create_directory(path, "Processed_.txt_Files", ".txt")
    log_directory_name = create_directory(path, ".log Files", ".log")

    # Create new text files to store data, or append to existing files if they already exist
    print("\nOne moment please...")
    file_name_1 = path + "\\" + "Defib_Shock_Extracted_Data.txt"
    file_name_2 = path + "\\" + "Defib_Shock_Case_Tracker.txt"
    totals_file = open(file_name_1, "a+")
    case_file = open(file_name_2, "a+")
    # Iterate over every relevant file in the folder and manipulate them
    for j in range(0, len(user_list)):
        file_name = user_list[j]
        file_path = path + "\\" + user_list[j]
        # Move any .log files to specified directory
        if file_name[len(file_name) - 4: len(file_name)] == ".log":
            shutil.move(file_path, log_directory_name + "\\" + user_list[j])
            continue

        # Remove extra characters from the file name to paste into the master data file
        updated_name_format = remove_extra_characters(file_name)

        # Read the selected .txt file line by line and copy the Defib Shock times to a new file
        file = open(file_path, "r")
        number_of_shocks = 0
        for line in file:
            line = line.lstrip()
            try:
                int_check = int(line[0])  # Use this statement to check if the leading character is an integer
                defib_shock_position = line.find("DEFIB SHOCK")
                if defib_shock_position != -1:
                    # Save name of line item
                    defib_shock = line[defib_shock_position: defib_shock_position + 11]
                    # Save time of element
                    front_bracket = line.find("[")
                    rear_bracket = line.find("]")
                    line = line[front_bracket + 1: rear_bracket]
                    line = line.lstrip()
                    try:
                        time = float(line)
                        totals_file.write(defib_shock + "  |  " + str(time) + "  |  " + updated_name_format + "\n")
                        number_of_shocks += 1
                    except ValueError:
                        pass
            except IndexError:
                continue
            except ValueError:
                continue
        file.close()

        # Write overall case details to the case .txt file
        shock_flag = "N"
        if number_of_shocks > 0:
            shock_flag = "Y"
        case_file.write(updated_name_format + "  |  " + shock_flag + "  |  " + str(number_of_shocks) + "\n")

        # Move current .txt to the directory for processed .txt files
        shutil.move(file_path, text_directory_name + "\\" + user_list[j])

    totals_file.close()
    case_file.close()

    print("\n.txt files manipulated successfully.")
    return file_name_1, file_name_2


# Create director for .txt and .log files if they haven't already been created
def create_directory(path, extension, file_type):
    directory_name = path + "\\" + extension
    try:
        os.mkdir(directory_name)
        print("\nSuccessfully created the directory %s to store processed %s files." % (directory_name, file_type))
    except OSError:
        print("\n%s files are stored in the directory %s." % (file_type, directory_name))
    finally:
        return directory_name


# Remove '.txt' and trailing digits from the end of the file name (this is what will be pasted in the file)
def remove_extra_characters(file_name):
    reversed_file_name = ""
    is_extra_digit = False
    # Reverse string to find last instance of '_'
    for k in range(len(file_name), 0, -1):
        reversed_file_name += file_name[k - 1]
    underscore_marker = reversed_file_name.find("_")
    # If '_' has a character (letter) to its left, it doesn't have extra digits
    if underscore_marker != -1:
        try:
            character_or_integer = int(file_name[len(file_name) - 1 - underscore_marker - 1])
            is_extra_digit = True
        except ValueError:
            is_extra_digit = False
        # If there are extra digits, remove them along with the file extension
        if is_extra_digit:
            text_to_remove = file_name[
                             len(file_name) - 1 - underscore_marker: len(file_name)]
            return file_name.replace(text_to_remove, "")
        else:
            return file_name.replace(".txt", "")
    else:
        return file_name.replace(".txt", "")


# Create a .csv file containing all of the data from the text file
def create_csv(path, title1, title2, title3):
    # Create the CSV file
    text_file = path
    csv_file = path.replace(".txt", ".csv")
    text_input = csv.reader(open(text_file), delimiter = "|")
    csv_output = csv.writer(open(csv_file, "w", newline = "\n"))
    csv_output.writerow([title1, title2, title3])
    csv_output.writerows(text_input)
    return csv_file


# Create an Excel file based on the .csv files, also contains basic statistics pertaining to defib shock data
def write_excel_remove_csv(data_path, case_path):
    # Write data to Excel file and format columns
    excel_file = data_path.replace(".csv", "_File.xlsx")
    excel_file = excel_file.replace("Extracted", "Master")
    writer = pd.ExcelWriter(excel_file, engine='xlsxwriter')
    df = pd.read_csv(data_path)
    df.to_excel(writer, sheet_name = "Defib Shock Extracted Data")
    df2 = pd.read_csv(case_path)
    df2.to_excel(writer, sheet_name = "Defib Shock Case Tracker")
    worksheet1 = writer.sheets["Defib Shock Extracted Data"]
    worksheet1.set_column('A:A', 4)
    worksheet1.set_column('B:D', 17)
    worksheet2 = writer.sheets["Defib Shock Case Tracker"]
    worksheet2.set_column('A:A', 4)
    worksheet2.set_column('B:D', 17)
    writer.save()
    os.remove(data_path)
    os.remove(case_path)
    writer.close()
    return excel_file


# Add Statistics to the master Excel file and eliminate white space from entries
def add_stats(excel_path):
    workbook = openpyxl.load_workbook(filename = excel_path)
    workbook.create_sheet("Defib Shock Statistics")
    number_of_cases = 0
    number_of_cases_with_shocks = 0
    number_of_shocks = []
    number_of_shocks_in_shock_cases = []
    # Iterate over every element in the workbook
    for i in range(0, 2):
        worksheet = workbook.worksheets[i]
        for j in range(1, worksheet.max_column + 1):
            for k in range(2, worksheet.max_row + 1):
                try:
                    worksheet.cell(row = k, column = j).value = float(str(worksheet.cell(row = k, column = j).
                                                                          value).strip())
                except ValueError:
                    worksheet.cell(row = k, column = j).value = str(worksheet.cell(row = k, column = j).value).strip()
                # If statement to line up with column in case tracker sheet recording 'Y/N', but not the title row
                if i == 1 and j == 3 and k != 1:
                    # Count total number of cases
                    number_of_cases += 1
                    # Count total number of cases with shocks occuring
                    if worksheet.cell(row = k, column = j).value == 'Y':
                        number_of_cases_with_shocks += 1
                # Count the total number of shocks of all cases combined
                if i == 1 and j == 4 and k != 1:
                    case_shock_count = worksheet.cell(row = k, column = j).value
                    number_of_shocks.append(case_shock_count)
                    if case_shock_count > 0:
                        number_of_shocks_in_shock_cases.append(case_shock_count)

    # Calculate remaining statistics
    shock_count = 0
    for p in range(0, len(number_of_shocks)):
        shock_count += number_of_shocks[p]
    shock_percent = float(100 * (number_of_cases_with_shocks / number_of_cases))
    shock_mean = float(shock_count/number_of_cases)
    shock_mean2 = float(shock_count/number_of_cases_with_shocks)
    shock_std_dev = float(statistics.stdev(number_of_shocks, shock_mean))
    shock_std_dev2 = float(statistics.stdev(number_of_shocks_in_shock_cases, shock_mean2))

    # Paste statistics into workbook
    worksheet = workbook.worksheets[2]
    worksheet.cell(row=1, column=1).value = "Total Number of Cases:"
    worksheet.cell(row=2, column=1).value = "Number of Cases with Shocks:"
    worksheet.cell(row=3, column=1).value = "Percent of Cases with a Shock (%):"
    worksheet.cell(row=4, column=1).value = "Total Number of Shocks:"
    worksheet.cell(row=5, column=1).value = "Average Number of Shocks per Case:"
    worksheet.cell(row=6, column=1).value = "Average Number of Shocks in Cases with Shocks:"
    worksheet.cell(row=7, column=1).value = "Shock Standard Deviation:"
    worksheet.cell(row=8, column=1).value = "Shock Standard Deviation in Cases with Shocks:"
    worksheet.cell(row=1, column=2).value = number_of_cases
    worksheet.cell(row=2, column=2).value = number_of_cases_with_shocks
    worksheet.cell(row=3, column=2).value = shock_percent
    worksheet.cell(row=4, column=2).value = shock_count
    worksheet.cell(row=5, column=2).value = shock_mean
    worksheet.cell(row=6, column=2).value = shock_mean2
    worksheet.cell(row=7, column=2).value = shock_std_dev
    worksheet.cell(row=8, column=2).value = shock_std_dev2
    # Change title column dimensions
    worksheet.column_dimensions['A'].width = 44
    # Make all titles bold in stats sheet
    for m in range(1, 9):
        worksheet.cell(row = m, column = 1).font = Font(bold = True)
    workbook.save(excel_path)


data_extraction_file, case_file = run_data_manipulation(
    r"C:\Users\mnarcisi\Documents\Mike\Scientific Affairs\Data_Initil_Testing")
if data_extraction_file != "N/A" and case_file != "N/A":
    data_csv_file = create_csv(data_extraction_file, "Element Name", "Time (sec)", "File Name")
    case_csv_file = create_csv(case_file, "File Name", "Shock (Y/N)", "Number of Shocks")
    excel_file = write_excel_remove_csv(data_csv_file, case_csv_file)
    add_stats(excel_file)