import os
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Color, PatternFill, Border, colors
import time
import math
import statistics


# Function used to link data files for compressions and shocks and find pre, post, and perri shock times
def find_pause_times(path):
    file_list = os.listdir(path)

    # If the master pause data file is open, abort the script and ask the user to close it to avoid an exception
    if "Pause_Master_Data_File.xlsx" in file_list:
        try:
            os.rename(path + "\\" + "Pause_Master_Data_File.xlsx", path + "\\" + "Pause_Master_Data_File.xlsx")
        except PermissionError:
            print("\nMaster Pause Data file is open.  Aborting.  Please close and re-run script.")
            return

    # Create master pause data file.  Will overwrite and create a new file if one previously exists.
    master_workbook = openpyxl.Workbook()
    master_workbook.create_sheet("Data with Errors")

    for x in range(0, 2):
        worksheet = master_workbook.worksheets[x]
        worksheet.cell(row=1, column=1).value = "Case ID"
        worksheet.cell(row=1, column=2).value = "CPR Start"
        worksheet.cell(row=1, column=3).value = "Compression Before Shock"
        worksheet.cell(row=1, column=4).value = "Shock"
        worksheet.cell(row=1, column=5).value = "Compression After Shock"
        worksheet.cell(row=1, column=6).value = "CPR End"
        worksheet.cell(row=1, column=7).value = "Pre-Shock Time"
        worksheet.cell(row=1, column=8).value = "Post-Shock Time"
        worksheet.cell(row=1, column=9).value = "Perri-Shock Time"
        for p in range(1, 10):
            worksheet.cell(row=1, column=p).font = Font(bold=True)
            worksheet.column_dimensions[get_column_letter(p)].width = 18
        worksheet.column_dimensions[get_column_letter(3)].width = 26
        worksheet.column_dimensions[get_column_letter(5)].width = 26
    master_row = 2  # Current row to paste data in the master sheet
    error_row = 2  # Current row to paste data in the error sheet

    # Initialize directories for missing or incorrect file names
    no_text_directory = []
    no_excel_directory = []
    no_text_or_excel_directory = []
    check_missing_cpr_cases = []

    # Initialize Data Collection Directories
    pre_shock_times = []
    post_shock_times = []
    perri_shock_times = []

    # Initialize Data Workbooks
    cpr_period_file = openpyxl.load_workbook("CPR_ROSC_PERIODS.xlsx")
    cpr_sheet = cpr_period_file.active
    shock_master_data_file = openpyxl.load_workbook(
        filename=r"C:\Users\mnarcisi\Documents\Mike\Scientific Affairs\Data_Group_2\Defib_Shock_Master_Data_File.xlsx")
    shock_sheet = shock_master_data_file.active

    # Cycle through all case files with known CPR periods and ensure data exists for each case
    text_file_directory = os.listdir(
        r"C:\Users\mnarcisi\Documents\Mike\Scientific Affairs\Data_Group_2\Processed_.txt_Files")
    # Truncate file extension from text file names in list
    for i in range(0, len(text_file_directory)):
        text_file_directory[i] = remove_extra_characters(text_file_directory[i], "Text")
        check_missing_cpr_cases.append(text_file_directory[i])
    excel_file_directory = os.listdir(
        r"C:\Users\mnarcisi\Documents\Mike\Scientific Affairs\Excel_File_Testing\Processed_Excel_Files")
    # Truncate extra characters from end of Excel file names in list
    for j in range(0, len(excel_file_directory)):
        excel_file_directory[j] = remove_extra_characters(excel_file_directory[j], "Excel")
        if excel_file_directory[j] not in check_missing_cpr_cases:
            check_missing_cpr_cases.append(excel_file_directory[j])
    # Cycle through each case in CPR period sheet
    for k in range(2, cpr_sheet.max_row):
        # Record case file name from CPR period data sheet
        case = cpr_sheet.cell(row=k, column=1).value
        # Check if current case is in list of text and Excel files.  If so, remove it.  (This is to check for missing
        # CPR period cases)
        if case in check_missing_cpr_cases:
            check_missing_cpr_cases.remove(case)
        # Check if current case file matches any stored text files containing defib shock data
        if case in text_file_directory:
            # Check if current case file matches any stored Excel files containing compression data
            if case in excel_file_directory:
                # Check that all compression and shock times fall within a CPR window
                try:
                    cpr_start = int(cpr_sheet.cell(row=k, column=2).value)
                except TypeError:
                    continue  # Skip rows that aren't CPR time rows
                try:
                    cpr_end = int(cpr_sheet.cell(row=k, column=4).value)
                    if cpr_end == 0:
                        continue  # Skip rows that aren't CPR time rows
                except TypeError:
                    continue  # Skip rows that aren't CPR time rows
                # Search Master Shock Date File for Shocks associated with the current case
                for m in range(2, shock_sheet.max_row):
                    # If shock data found, search for associated compression file
                    if case == shock_sheet.cell(row=m, column=4).value:
                        shock_time = int(shock_sheet.cell(row=m, column=3).value * 1000)
                        try:
                            compression_case_file = openpyxl.load_workbook(
                                filename=r"C:\Users\mnarcisi\Documents\Mike\Scientific Affairs\Excel_File_Testing\Processed_Excel_Files" + "\\" + case + ".xlsx")
                            compression_sheet = compression_case_file.active
                        except FileNotFoundError:
                            try:
                                compression_case_file = openpyxl.load_workbook(
                                    filename=r"C:\Users\mnarcisi\Documents\Mike\Scientific Affairs\Excel_File_Testing\Processed_Excel_Files" + "\\" + case + "_01.xlsx")
                                compression_sheet = compression_case_file.active
                            except FileNotFoundError:
                                try:
                                    compression_case_file = openpyxl.load_workbook(
                                        filename=r"C:\Users\mnarcisi\Documents\Mike\Scientific Affairs\Excel_File_Testing\Processed_Excel_Files" + "\\" + case + "_02.xlsx")
                                    compression_sheet = compression_case_file.active
                                except FileNotFoundError:
                                    continue

                        # Search compression file for compressions performed surrounding the shock
                        max_row = compression_sheet.max_row
                        pre_compression_time = 0
                        post_compression_time = 0
                        for n in range(2, max_row + 1):
                            if int(compression_sheet.cell(row=n, column=4).value) > shock_time:
                                post_compression_time = int(compression_sheet.cell(row=n, column=4).value)
                                try:
                                    pre_compression_time = int(compression_sheet.cell(row=n-1, column=4).value)
                                except ValueError:
                                    pre_compression_time = 0
                                finally:
                                    break
                        # If no more compressions after shock, make pre-shock = last compression, and post-shock =
                        # the end of the CPR period (if this doesn't work, there are no compressions, so we skip)
                        if post_compression_time == 0:
                            try:
                                pre_compression_time = int(compression_sheet.cell(row=max_row, column=4).value)
                                post_compression_time = cpr_end
                            except ValueError:
                                continue

                        pre_shock_time = shock_time - pre_compression_time
                        post_shock_time = post_compression_time - shock_time
                        perri_shock_time = post_compression_time - pre_compression_time

                        # Flag rows with errors and remove from data, posting in a separate tab
                        if shock_time < cpr_start or shock_time > cpr_end:
                            error_flag = 1
                        elif (cpr_end < pre_compression_time) or (cpr_end < post_compression_time) or \
                            (cpr_start > pre_compression_time) or (cpr_start > post_compression_time):
                            error_flag = 1
                        elif (post_compression_time < shock_time) or (post_compression_time == 0) or \
                                (pre_compression_time > shock_time):
                            error_flag = 1
                        else:
                            error_flag = 0

                        # Record relevant values to the correct data sheet
                        if error_flag == 0:
                            worksheet = master_workbook.worksheets[0]
                            row = master_row
                        else:
                            worksheet = master_workbook.worksheets[1]
                            row = error_row

                        worksheet.cell(row=row, column=1).value = case
                        worksheet.cell(row=row, column=2).value = cpr_start
                        worksheet.cell(row=row, column=3).value = pre_compression_time
                        worksheet.cell(row=row, column=4).value = shock_time
                        worksheet.cell(row=row, column=5).value = post_compression_time
                        worksheet.cell(row=row, column=6).value = cpr_end
                        worksheet.cell(row=row, column=7).value = pre_shock_time
                        worksheet.cell(row=row, column=8).value = post_shock_time
                        worksheet.cell(row=row, column=9).value = perri_shock_time

                        # Mark cells with data errors
                        incorrect_cpr_end_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
                        incorrect_shock_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
                        incorrect_compression_fill = PatternFill(start_color='FF7900', end_color='FF7900', fill_type='solid')
                        # Highlight if data shows the shock was performed before or after the CPR period (RED)
                        if error_flag == 1:
                            if shock_time < cpr_start:
                                worksheet[get_column_letter(2) + str(row)].fill = incorrect_shock_fill
                                worksheet[get_column_letter(4) + str(row)].fill = incorrect_shock_fill
                            elif shock_time > cpr_end:
                                worksheet[get_column_letter(6) + str(row)].fill = incorrect_shock_fill
                                worksheet[get_column_letter(4) + str(row)].fill = incorrect_shock_fill
                        # Highlight if data shows the compression was performed before or after the CPR period
                        # (but not the shock) (ORANGE)
                            if cpr_end < pre_compression_time:
                                worksheet[get_column_letter(3) + str(row)].fill = incorrect_compression_fill
                                if worksheet[get_column_letter(6) + str(row)].fill != incorrect_shock_fill:
                                    worksheet[get_column_letter(6) + str(row)].fill = incorrect_compression_fill
                            elif cpr_end < post_compression_time:
                                worksheet[get_column_letter(5) + str(row)].fill = incorrect_compression_fill
                                if worksheet[get_column_letter(6) + str(row)].fill != incorrect_shock_fill:
                                    worksheet[get_column_letter(6) + str(row)].fill = incorrect_compression_fill
                            elif cpr_start > pre_compression_time:
                                worksheet[get_column_letter(3) + str(row)].fill = incorrect_compression_fill
                                if worksheet[get_column_letter(2) + str(row)].fill != incorrect_shock_fill:
                                    worksheet[get_column_letter(2) + str(row)].fill = incorrect_compression_fill
                            else:
                                if worksheet[get_column_letter(2) + str(row)].fill != incorrect_shock_fill:
                                    worksheet[get_column_letter(2) + str(row)].fill = incorrect_compression_fill
                                worksheet[get_column_letter(5) + str(row)].fill = incorrect_compression_fill
                        # Highlights if compression times are incorrect or missing (GREEN)
                            if pre_compression_time > shock_time:
                                if worksheet[get_column_letter(4) + str(row)].fill != incorrect_shock_fill:
                                    worksheet[get_column_letter(4) + str(row)].fill = incorrect_cpr_end_fill
                                if worksheet[get_column_letter(3) + str(row)].fill != incorrect_compression_fill:
                                    worksheet[get_column_letter(3) + str(row)].fill = incorrect_cpr_end_fill
                            elif post_compression_time < shock_time or (post_compression_time == 0):
                                if worksheet[get_column_letter(4) + str(row)].fill != incorrect_shock_fill:
                                    worksheet[get_column_letter(4) + str(row)].fill = incorrect_cpr_end_fill
                                if worksheet[get_column_letter(5) + str(row)].fill != incorrect_compression_fill:
                                    worksheet[get_column_letter(5) + str(row)].fill = incorrect_cpr_end_fill

                        # Continue to the next row of the proper data sheet, record times if master sheet
                        if error_flag == 0:
                            master_row += 1
                            pre_shock_times.append(pre_shock_time)
                            post_shock_times.append(post_shock_time)
                            perri_shock_times.append(perri_shock_time)
                        else:
                            error_row += 1

            # If case has a text file but not an Excel file
            else:
                no_excel_directory.append(case)
        else:
            # If case doesn't have a text file but has an Excel file
            if case in excel_file_directory:
                no_text_directory.append(case)
            # If case doesn't have a text or an Excel file
            else:
                no_text_or_excel_directory.append(case)

    # Add second workbook tab for Overall Statistics (if there are any to calculate)
    if len(pre_shock_times) > 1:
        master_workbook.create_sheet("Pause Statistics")
        data_sheet = master_workbook.worksheets[2]

        data_sheet.cell(row=1, column=1).value = "Milliseconds"

        data_sheet.cell(row=3, column=1).value = "Pre-Shock Time"
        data_sheet.cell(row=4, column=1).value = "Post-Shock Time"
        data_sheet.cell(row=5, column=1).value = "Perri-Shock Time"

        data_sheet.cell(row=2, column=2).value = "Mean"
        data_sheet.cell(row=2, column=3).value = "Minimum"
        data_sheet.cell(row=2, column=4).value = "Maximum"
        data_sheet.cell(row=2, column=5).value = "Standard Deviation"
        data_sheet.cell(row=2, column=6).value = "Variance"
        data_sheet.cell(row=2, column=7).value = "Median"
        data_sheet.cell(row=2, column=8).value = "Interquartile Range"
        data_sheet.cell(row=2, column=9).value = "Standard Error"
        for q in range(1, 10):
            data_sheet.cell(row=2, column=q).font = Font(bold=True)
            data_sheet.column_dimensions[get_column_letter(q)].width = 18
        data_sheet.column_dimensions[get_column_letter(5)].width = 26
        data_sheet.column_dimensions[get_column_letter(8)].width = 26
        data_sheet.column_dimensions[get_column_letter(9)].width = 26
        data_sheet.cell(row=1, column=1).font = Font(bold=True)
        data_sheet.cell(row=3, column=1).font = Font(bold=True)
        data_sheet.cell(row=4, column=1).font = Font(bold=True)
        data_sheet.cell(row=5, column=1).font = Font(bold=True)

        # Gather necessary statistics from Master Sheet
        pre_shock_times.sort()
        post_shock_times.sort()
        perri_shock_times.sort()
        pre_shock_mean = float(sum(pre_shock_times)/len(pre_shock_times))
        post_shock_mean = float(sum(post_shock_times)/len(post_shock_times))
        perri_shock_mean = float(sum(perri_shock_times)/len(perri_shock_times))
        try:
            pre_shock_std_dev = statistics.stdev(pre_shock_times, pre_shock_mean)
        except statistics.StatisticsError:
            pre_shock_std_dev = 0
        try:
            post_shock_std_dev = statistics.stdev(post_shock_times, post_shock_mean)
        except statistics.StatisticsError:
            post_shock_std_dev = 0
        try:
            perri_shock_std_dev = statistics.stdev(perri_shock_times, perri_shock_mean)
        except statistics.StatisticsError:
            perri_shock_std_dev = 0
        pre_shock_median = get_median_value(pre_shock_times)
        post_shock_median = get_median_value(post_shock_times)
        perri_shock_median = get_median_value(perri_shock_times)

        # Mean Values
        data_sheet.cell(row=3, column=2).value = pre_shock_mean
        data_sheet.cell(row=4, column=2).value = post_shock_mean
        data_sheet.cell(row=5, column=2).value = perri_shock_mean

        # Maximum and Minimum Values
        data_sheet.cell(row=3, column=3).value = pre_shock_times[0]
        data_sheet.cell(row=4, column=3).value = post_shock_times[0]
        data_sheet.cell(row=5, column=3).value = perri_shock_times[0]
        data_sheet.cell(row=3, column=4).value = pre_shock_times[len(pre_shock_times) - 1]
        data_sheet.cell(row=4, column=4).value = post_shock_times[len(post_shock_times) - 1]
        data_sheet.cell(row=5, column=4).value = perri_shock_times[len(perri_shock_times) - 1]

        # Standard Deviation and Variance Values
        data_sheet.cell(row=3, column=5).value = pre_shock_std_dev
        data_sheet.cell(row=4, column=5).value = post_shock_std_dev
        data_sheet.cell(row=5, column=5).value = perri_shock_std_dev
        data_sheet.cell(row=3, column=6).value = math.pow(pre_shock_std_dev, 2)
        data_sheet.cell(row=4, column=6).value = math.pow(post_shock_std_dev, 2)
        data_sheet.cell(row=5, column=6).value = math.pow(perri_shock_std_dev, 2)

        # Median Values
        data_sheet.cell(row=3, column=7).value = pre_shock_median
        data_sheet.cell(row=4, column=7).value = post_shock_median
        data_sheet.cell(row=5, column=7).value = perri_shock_median

        # Interquartile Range Values
        data_sheet.cell(row=3, column=8).value = get_interquartile_range_values(pre_shock_times, pre_shock_median)
        data_sheet.cell(row=4, column=8).value = get_interquartile_range_values(post_shock_times, post_shock_median)
        data_sheet.cell(row=5, column=8).value = get_interquartile_range_values(perri_shock_times, perri_shock_median)

        # Standard Error Values
        data_sheet.cell(row=3, column=9).value = pre_shock_std_dev / math.sqrt(len(pre_shock_times))
        data_sheet.cell(row=4, column=9).value = post_shock_std_dev / math.sqrt(len(post_shock_times))
        data_sheet.cell(row=5, column=9).value = perri_shock_std_dev / math.sqrt(len(perri_shock_times))

        # Add section for all data values converted to seconds
        data_sheet.cell(row=7, column=1).value = "Seconds"

        data_sheet.cell(row=9, column=1).value = "Pre-Shock Time"
        data_sheet.cell(row=10, column=1).value = "Post-Shock Time"
        data_sheet.cell(row=11, column=1).value = "Perri-Shock Time"

        data_sheet.cell(row=8, column=2).value = "Mean"
        data_sheet.cell(row=8, column=3).value = "Minimum"
        data_sheet.cell(row=8, column=4).value = "Maximum"
        data_sheet.cell(row=8, column=5).value = "Standard Deviation"
        data_sheet.cell(row=8, column=6).value = "Variance"
        data_sheet.cell(row=8, column=7).value = "Median"
        data_sheet.cell(row=8, column=8).value = "Interquartile Range"
        data_sheet.cell(row=8, column=9).value = "Standard Error"
        for r in range(1, 10):
            data_sheet.cell(row=8, column=r).font = Font(bold=True)
        data_sheet.cell(row=7, column=1).font = Font(bold=True)
        data_sheet.cell(row=9, column=1).font = Font(bold=True)
        data_sheet.cell(row=10, column=1).font = Font(bold=True)
        data_sheet.cell(row=11, column=1).font = Font(bold=True)

        # Mean Values
        data_sheet.cell(row=9, column=2).value = data_sheet.cell(row=3, column=2).value / 1000
        data_sheet.cell(row=10, column=2).value = data_sheet.cell(row=4, column=2).value / 1000
        data_sheet.cell(row=11, column=2).value = data_sheet.cell(row=5, column=2).value / 1000

        # Maximum and Minimum Values
        data_sheet.cell(row=9, column=3).value = data_sheet.cell(row=3, column=3).value / 1000
        data_sheet.cell(row=10, column=3).value = data_sheet.cell(row=4, column=3).value / 1000
        data_sheet.cell(row=11, column=3).value = data_sheet.cell(row=5, column=3).value / 1000
        data_sheet.cell(row=9, column=4).value = data_sheet.cell(row=3, column=4).value / 1000
        data_sheet.cell(row=10, column=4).value = data_sheet.cell(row=4, column=4).value / 1000
        data_sheet.cell(row=11, column=4).value = data_sheet.cell(row=5, column=4).value / 1000

        # Standard Deviation and Variance Values
        data_sheet.cell(row=9, column=5).value = data_sheet.cell(row=3, column=5).value / 1000
        data_sheet.cell(row=10, column=5).value = data_sheet.cell(row=4, column=5).value / 1000
        data_sheet.cell(row=11, column=5).value = data_sheet.cell(row=5, column=5).value / 1000
        data_sheet.cell(row=9, column=6).value = math.pow(data_sheet.cell(row=3, column=5).value / 1000, 2)
        data_sheet.cell(row=10, column=6).value = math.pow(data_sheet.cell(row=4, column=5).value / 1000, 2)
        data_sheet.cell(row=11, column=6).value = math.pow(data_sheet.cell(row=5, column=5).value / 1000, 2)

        # Median Values
        data_sheet.cell(row=9, column=7).value = data_sheet.cell(row=3, column=7).value / 1000
        data_sheet.cell(row=10, column=7).value = data_sheet.cell(row=4, column=7).value / 1000
        data_sheet.cell(row=11, column=7).value = data_sheet.cell(row=5, column=7).value / 1000

        # Interquartile Range Values
        data_sheet.cell(row=9, column=8).value = data_sheet.cell(row=3, column=8).value / 1000
        data_sheet.cell(row=10, column=8).value = data_sheet.cell(row=4, column=8).value / 1000
        data_sheet.cell(row=11, column=8).value = data_sheet.cell(row=5, column=8).value / 1000

        # Standard Error Values
        data_sheet.cell(row=9, column=9).value = (data_sheet.cell(row=3, column=5).value / 1000) / math.sqrt(
            len(pre_shock_times))
        data_sheet.cell(row=10, column=9).value = (data_sheet.cell(row=4, column=5).value / 1000) / math.sqrt(
            len(post_shock_times))
        data_sheet.cell(row=11, column=9).value = (data_sheet.cell(row=5, column=5).value / 1000) / math.sqrt(
            len(perri_shock_times))

    # Add a sheet to the workbook which displays missing files
    master_workbook.create_sheet("Missing Files")
    missing_file_sheet = master_workbook.worksheets[3]

    missing_file_sheet.cell(row=1, column=1).value = "Cases Missing Shock File Only"
    missing_file_sheet.cell(row=1, column=2).value = "Cases Missing Compression File Only"
    missing_file_sheet.cell(row=1, column=3).value = "Cases Missing Shock and Compression Files"
    missing_file_sheet.cell(row=1, column=4).value = "Cases Missing CPR Window Data"
    for p in range(1, 5):
        missing_file_sheet.cell(row=1, column=p).font = Font(bold=True)
        missing_file_sheet.column_dimensions[get_column_letter(p)].width = 40

    for r1 in range(2, len(no_text_directory) + 1):
        missing_file_sheet.cell(row=r1, column=1).value = no_text_directory[r1 - 2]
    for r2 in range(2, len(no_excel_directory) + 1):
        missing_file_sheet.cell(row=r2, column=2).value = no_excel_directory[r2 - 2]
    for r3 in range(2, len(no_text_or_excel_directory) + 1):
        missing_file_sheet.cell(row=r3, column=3).value = no_text_or_excel_directory[r3 - 2]
    for r4 in range(2, len(check_missing_cpr_cases) + 1):
        missing_file_sheet.cell(row=r4, column=4).value = check_missing_cpr_cases[r4 - 2]

    # Save Master Data File
    save_path = path + "\\" + "Pause_Master_Data_File.xlsx"
    master_workbook.save(filename=save_path)
    master_workbook.close()
    print("\nPause Data Workbook saved to " + save_path + ".")


# Remove un-needed characters from the end of the file name
def remove_extra_characters(file_name, file_type):
    reversed_file_name = ""
    # Reverse string to find last instance of '_'
    for k in range(len(file_name), 0, -1):
        reversed_file_name += file_name[k - 1]
    underscore_marker = reversed_file_name.find("_")
    # If no '_', just eliminate file extension
    if underscore_marker == -1:
        # Remove only 4 characters if it is a text file, and 5 if it is an Excel file
        if file_type == "Text":
            return file_name.replace(file_name[len(file_name) - 4: len(file_name)], "")
        else:
            return file_name.replace(file_name[len(file_name) - 5: len(file_name)], "")
    # If '_' has a character to its left, it doesn't have extra digits
    else:
        try:
            # If there are extra digits, remove them along with the file extension
            char_or_int = int(file_name[len(file_name) - 1 - underscore_marker - 1])
            return file_name.replace(file_name[len(file_name) - 1 - underscore_marker: len(file_name)], "")
        except ValueError:
            if file_type == "Text":
                return file_name.replace(file_name[len(file_name) - 4: len(file_name)], "")
            else:
                return file_name.replace(file_name[len(file_name) - 5: len(file_name)], "")


# Calculate Median Values
def get_median_value(data):
    # Process is different depending on if the length of the data is even or odd
    if len(data) % 2 == 0:
        upper_value = data[math.ceil(len(data)/2)]
        lower_value = data[math.ceil(len(data)/2) - 1]
        return float((lower_value + upper_value) / 2)
    else:
        return data[math.ceil(len(data)/2) - 1]


# Calculate Interquartile Range Values
def get_interquartile_range_values(data, median):
    # Split data between lower and upper half
    lower_half = []
    upper_half = []
    for value in data:
        if value < median:
            lower_half.append(value)
        elif value > median:
            upper_half.append(value)
        else:
            pass  # If value = median, don't include it in either list
    q1_range = get_median_value(lower_half)
    q3_range = get_median_value(upper_half)
    return q3_range - q1_range


start_time = time.time()
find_pause_times(r"C:\Users\mnarcisi\Documents\Mike\Scientific Affairs")
# Print time taken to run script
elapsed_time = time.time() - start_time
print("\nTotal time to run script: " + str(round(elapsed_time, 3)) + " seconds.")