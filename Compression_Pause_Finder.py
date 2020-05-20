import os
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Color, PatternFill, Border, colors
import time
import math
import statistics

# NOTE: This script takes about 30 minutes to run for 2000 files


# Function used to find the Pauses between compressions during each CPR incident, ensuring not to include artifact
def find_compression_pauses(path):
    file_list = os.listdir(path)
    new_compression_file_path = r"C:\Users\mnarcisi\Documents\Mike\Scientific Affairs\Updated_Compression_Files"
    new_compression_list = os.listdir(new_compression_file_path)

    # If the master compression pause data file is open, abort script and ask user to close it to avoid an exception
    if "Compression_Pause_Master_Data_File.xlsx" in file_list:
        try:
            os.rename(path + "\\" + "Compression_Pause_Master_Data_File.xlsx", path + "\\" +
                      "Compression_Pause_Master_Data_File.xlsx")
        except PermissionError:
            print("\nMaster Compression Pause Data file is open.  Aborting.  Please close and re-run script.")
            return

    # Create master compression pause data file.  Will overwrite and create a new file if one previously exists.
    master_workbook = openpyxl.Workbook()
    master_worksheet = master_workbook.active
    master_worksheet.title = "Case Tracker (Milliseconds)"
    master_workbook.create_sheet("Case Tracker (Seconds)")
    seconds_sheet = master_workbook.worksheets[1]

    for a in range(0, 2):
        worksheet = master_workbook.worksheets[a]
        worksheet.cell(row=1, column=1).value = "Case ID"
        worksheet.cell(row=1, column=2).value = "Total Compressions"
        worksheet.cell(row=1, column=3).value = "Total Compression Periods"
        worksheet.cell(row=1, column=4).value = "Total Pauses"
        worksheet.cell(row=1, column=5).value = "Compression Period Mean"
        worksheet.cell(row=1, column=6).value = "Compression Period Minimum"
        worksheet.cell(row=1, column=7).value = "Compression Period Maximum"
        worksheet.cell(row=1, column=8).value = "Compression Period Standard Deviation"
        worksheet.cell(row=1, column=9).value = "Compression Period Variance"
        worksheet.cell(row=1, column=10).value = "Compression Period Median"
        worksheet.cell(row=1, column=11).value = "Compression Period Interquartile Range"
        worksheet.cell(row=1, column=12).value = "Compression Period Standard Error"
        worksheet.cell(row=1, column=13).value = "Pause Mean"
        worksheet.cell(row=1, column=14).value = "Pause Minimum"
        worksheet.cell(row=1, column=15).value = "Pause Maximum"
        worksheet.cell(row=1, column=16).value = "Pause Standard Deviation"
        worksheet.cell(row=1, column=17).value = "Pause Variance"
        worksheet.cell(row=1, column=18).value = "Pause Median"
        worksheet.cell(row=1, column=19).value = "Pause Interquartile Range"
        worksheet.cell(row=1, column=20).value = "Pause Standard Error"
        for p in range(1, 21):
            worksheet.cell(row=1, column=p).font = Font(bold=True)
            worksheet.column_dimensions[get_column_letter(p)].width = 18
        worksheet.column_dimensions[get_column_letter(3)].width = 26
        worksheet.column_dimensions[get_column_letter(5)].width = 28
        worksheet.column_dimensions[get_column_letter(6)].width = 30
        worksheet.column_dimensions[get_column_letter(7)].width = 30
        worksheet.column_dimensions[get_column_letter(8)].width = 38
        worksheet.column_dimensions[get_column_letter(9)].width = 30
        worksheet.column_dimensions[get_column_letter(10)].width = 30
        worksheet.column_dimensions[get_column_letter(11)].width = 38
        worksheet.column_dimensions[get_column_letter(12)].width = 36
        worksheet.column_dimensions[get_column_letter(16)].width = 30
        worksheet.column_dimensions[get_column_letter(19)].width = 31
        worksheet.column_dimensions[get_column_letter(20)].width = 23

    # Initialize list variables
    cpr_period_list = []
    final_compression_file_list = []
    in_cpr_not_excel = []
    in_excel_not_cpr = []
    red_color = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    grey_color = PatternFill(start_color='999999', end_color='999999', fill_type='solid')
    yellow_color = PatternFill(start_color='FFFF33', end_color='454545', fill_type='solid')
    dark_green_color = PatternFill(start_color='005500', end_color='454545', fill_type='solid')

    # Fill list variables with Case ID numbers to later match up
    cpr_period_file = openpyxl.load_workbook("Clean_CPR_Periods.xlsx")
    cpr_sheet = cpr_period_file.active
    for i in range(2, cpr_sheet.max_row + 1):
        cpr_period_list.append(cpr_sheet.cell(row=i, column=1).value)

    # Loop through all Compression data files that have CPR periods, and output an updated Compression file for each
    original_compression_file_directory = \
        r"C:\Users\mnarcisi\Documents\Mike\Scientific Affairs\Excel_File_Testing\Processed_Excel_Files"
    compression_case_list = os.listdir(original_compression_file_directory)

    for j in range(0, len(compression_case_list)):
        case = remove_extra_characters_2(compression_case_list[j])
        final_compression_file_list.append(case)
        if case in cpr_period_list:
            # If the current compression pause data file is open, skip it and notify user
            permission_flag = True
            if case + ".xlsx" in new_compression_list:
                try:
                    os.rename(new_compression_file_path + "\\" + case + ".xlsx",
                              new_compression_file_path + "\\" + case + ".xlsx")
                except PermissionError:
                    print("\nCompression Case file " + case + ".xlsx" +
                          " is open.  Skipping and not adding numbers to master file.")
                    permission_flag = False

            if permission_flag:
                # Record the current case's CPR Periods
                cpr_start = []
                cpr_end = []
                cpr_indicator_flag = False
                for m in range(2, cpr_sheet.max_row + 1):
                    if cpr_sheet.cell(row=m, column=1).value == case:
                        cpr_start.append(cpr_sheet.cell(row=m, column=2).value)
                        cpr_end.append(cpr_sheet.cell(row=m, column=3).value)
                        cpr_indicator_flag = True
                    elif cpr_indicator_flag:
                        break

                # Create new Compression Workbook
                new_compression_workbook = openpyxl.Workbook()
                compression_worksheet = new_compression_workbook.active
                compression_worksheet.title = "Compression Data"

                compression_worksheet.cell(row=1, column=10).value = "In_CPR_Period?"
                compression_worksheet.cell(row=1, column=11).value = "Comp_Period (ms)"
                compression_worksheet.cell(row=1, column=12).value = "Comp_Period (s)"
                compression_worksheet.cell(row=1, column=13).value = "Pause?_(n>2000ms)"
                compression_worksheet.cell(row=1, column=14).value = "Pause_Artifact?"
                for p in range(1, 15):
                    compression_worksheet.cell(row=1, column=p).font = Font(bold=True)
                    compression_worksheet.column_dimensions[get_column_letter(p)].width = 15
                compression_worksheet.column_dimensions[get_column_letter(7)].width = 21
                compression_worksheet.column_dimensions[get_column_letter(9)].width = 9
                compression_worksheet.column_dimensions[get_column_letter(11)].width = 18
                compression_worksheet.column_dimensions[get_column_letter(12)].width = 19
                compression_worksheet.column_dimensions[get_column_letter(13)].width = 21

                # Fill new Worksheet with Existing Information from the original workbook
                old_compression_file = openpyxl.load_workbook(
                    original_compression_file_directory + "\\" + compression_case_list[j])
                old_compression_worksheet = old_compression_file.active
                for k in range(1, old_compression_worksheet.max_row + 1):
                    compression_worksheet.cell(row=k, column=1).value = old_compression_worksheet.cell(
                        row=k, column=1).value
                    compression_worksheet.cell(row=k, column=2).value = old_compression_worksheet.cell(
                        row=k, column=2).value
                    compression_worksheet.cell(row=k, column=3).value = old_compression_worksheet.cell(
                        row=k, column=3).value
                    compression_worksheet.cell(row=k, column=4).value = old_compression_worksheet.cell(
                        row=k, column=4).value
                    compression_worksheet.cell(row=k, column=5).value = old_compression_worksheet.cell(
                        row=k, column=5).value
                    compression_worksheet.cell(row=k, column=6).value = old_compression_worksheet.cell(
                        row=k, column=6).value
                    compression_worksheet.cell(row=k, column=7).value = old_compression_worksheet.cell(
                        row=k, column=7).value
                    compression_worksheet.cell(row=k, column=8).value = old_compression_worksheet.cell(
                        row=k, column=8).value

                    # Create Sheet Separator Column
                    compression_worksheet[get_column_letter(9) + str(k)].fill = red_color

                    if k > 1:
                        # Insert True or False to show if Compression was during one of the CPR periods
                        comp_in_cpr = "FALSE"
                        for h in range(0, len(cpr_start)):
                            if cpr_start[h] <= compression_worksheet.cell(row=k, column=4).value <= cpr_end[h]:
                                comp_in_cpr = "TRUE"
                        compression_worksheet.cell(row=k, column=10).value = comp_in_cpr

                        # Insert the Compression Period (Time between current and previous compression)
                        comp_period = ""
                        if k == 2 and comp_in_cpr == "TRUE":  # First compression should always have a "N/A" comp period
                            comp_period = "N/A"
                        elif comp_in_cpr == "TRUE":
                            # If previous compression wasn't in CPR period, this compression has no CPR period
                            if compression_worksheet.cell(row=k - 1, column=10).value == "FALSE":
                                comp_period = "N/A"
                            else:
                                comp_period = compression_worksheet.cell(row=k, column=4).value - \
                                              compression_worksheet.cell(row=k - 1, column=4).value
                        else:  # Don't include compression period data for compressions done outside of the CPR period
                            for f in range(1, 15):
                                if f != 9:
                                    compression_worksheet[get_column_letter(f) + str(k)].fill = dark_green_color
                        compression_worksheet.cell(row=k, column=11).value = comp_period
                        try:
                            compression_worksheet.cell(row=k, column=12).value = comp_period / 1000
                        except TypeError:
                            if comp_period == "N/A":
                                compression_worksheet.cell(row=k, column=12).value = "N/A"
                            else:
                                compression_worksheet.cell(row=k, column=12).value = ""

                        # If compression period is a number, continue formulating other data
                        try:
                            comp_period = int(comp_period)  # Ensure compression period is a number
                            # Insert Pause Indicator
                            if comp_period > 2000:
                                pause = "TRUE"
                                for q in range(1, 15):
                                    if q != 9:
                                        compression_worksheet[get_column_letter(q) + str(k)].fill = yellow_color
                            else:
                                pause = "FALSE"
                            compression_worksheet.cell(row=k, column=13).value = pause

                            # Insert Artifact Indicator
                            if pause == "TRUE":
                                # Find the index of the previous pause
                                previous_pause_index = 2
                                # Find the index of the previous pause, or compression before previous non-CPR period
                                for n in range(k - 1, 1, -1):
                                    if compression_worksheet.cell(row=n, column=13).value == "TRUE" or \
                                            compression_worksheet.cell(row=n-1, column=10).value == "FALSE":
                                        previous_pause_index = n
                                        break
                                # Flag less than 3 compressions between pauses as artifact
                                if k - previous_pause_index < 4:
                                    for w in range(previous_pause_index, k):
                                        compression_worksheet.cell(row=w, column=14).value = "Artifact"
                                        for x in range(1, 15):
                                            if x != 9:
                                                compression_worksheet[get_column_letter(x) + str(w)].fill = grey_color
                                    compression_worksheet.cell(row=k, column=14).value = "Lead Pause"
                                    for q in range(1, 15):
                                        if q != 9:
                                            compression_worksheet[get_column_letter(q) + str(k)].fill = yellow_color
                        except ValueError:
                            if comp_period == "N/A":
                                compression_worksheet.cell(row=k, column=13).value = "FALSE"

                # Loop through worksheet to adjust artifact calculations
                artifact_index = 1
                for v in range(2, compression_worksheet.max_row + 1):
                    updated_comp_period = 0
                    if compression_worksheet.cell(row=v, column=14).value == "Artifact":
                        artifact_index += 1
                        # Remove compression period from all compressions labelled as artifact
                        compression_worksheet.cell(row=v, column=11).value = ""
                        compression_worksheet.cell(row=v, column=12).value = ""
                    if compression_worksheet.cell(row=v, column=14).value == "Lead Pause":
                        try:
                            updated_comp_period = compression_worksheet.cell(row=v, column=4).value - \
                                                  compression_worksheet.cell(
                                                      row=(v - artifact_index), column=4).value
                        except TypeError:
                            updated_comp_period = compression_worksheet.cell(row=v, column=4).value - 0
                        finally:
                            compression_worksheet.cell(row=v, column=11).value = updated_comp_period
                            compression_worksheet.cell(row=v, column=12).value = updated_comp_period / 1000
                            artifact_index = 1

                # Create second sheet with Statistics for the Individual Case
                new_compression_workbook.create_sheet("Case Compression Statistics")
                stats_sheet = new_compression_workbook.worksheets[1]

                stats_sheet.cell(row=1, column=1).value = "Case Number"

                stats_sheet.cell(row=3, column=1).value = "Compression Period Statistics"
                stats_sheet.cell(row=3, column=2).value = "Mean"
                stats_sheet.cell(row=3, column=3).value = "Minimum"
                stats_sheet.cell(row=3, column=4).value = "Maximum"
                stats_sheet.cell(row=3, column=5).value = "Standard Deviation"
                stats_sheet.cell(row=3, column=6).value = "Variance"
                stats_sheet.cell(row=3, column=7).value = "Median"
                stats_sheet.cell(row=3, column=8).value = "Interquartile Range"
                stats_sheet.cell(row=3, column=9).value = "Standard Error"

                stats_sheet.cell(row=4, column=1).value = "Milliseconds"
                stats_sheet.cell(row=5, column=1).value = "Seconds"

                stats_sheet.cell(row=7, column=1).value = "Pause Statistics"
                stats_sheet.cell(row=7, column=2).value = "Mean"
                stats_sheet.cell(row=7, column=3).value = "Minimum"
                stats_sheet.cell(row=7, column=4).value = "Maximum"
                stats_sheet.cell(row=7, column=5).value = "Standard Deviation"
                stats_sheet.cell(row=7, column=6).value = "Variance"
                stats_sheet.cell(row=7, column=7).value = "Median"
                stats_sheet.cell(row=7, column=8).value = "Interquartile Range"
                stats_sheet.cell(row=7, column=9).value = "Standard Error"

                stats_sheet.cell(row=8, column=1).value = "Milliseconds"
                stats_sheet.cell(row=9, column=1).value = "Seconds"

                stats_sheet.cell(row=11, column=1).value = "Total Compressions"
                stats_sheet.cell(row=12, column=1).value = "(Minus Artifact and Data outside of CPR Period)"

                stats_sheet.cell(row=14, column=1).value = "Total Compression Periods"
                stats_sheet.cell(row=15, column=1).value = "(Minus Artifact and Data outside of CPR Period)"

                stats_sheet.cell(row=17, column=1).value = "Total Pauses"
                stats_sheet.cell(row=18, column=1).value = "(Minus Artifact and Data outside of CPR Period)"

                stats_sheet.cell(row=1, column=1).font = Font(bold=True)
                for q in range(1, 13):
                    stats_sheet.cell(row=3, column=q).font = Font(bold=True)
                    stats_sheet.cell(row=7, column=q).font = Font(bold=True)
                    stats_sheet.column_dimensions[get_column_letter(q)].width = 18
                stats_sheet.column_dimensions[get_column_letter(1)].width = 38
                stats_sheet.column_dimensions[get_column_letter(5)].width = 23
                stats_sheet.column_dimensions[get_column_letter(8)].width = 24
                stats_sheet.column_dimensions[get_column_letter(9)].width = 20
                stats_sheet.cell(row=4, column=1).font = Font(bold=True)
                stats_sheet.cell(row=5, column=1).font = Font(bold=True)
                stats_sheet.cell(row=8, column=1).font = Font(bold=True)
                stats_sheet.cell(row=9, column=1).font = Font(bold=True)
                stats_sheet.cell(row=11, column=1).font = Font(bold=True)
                stats_sheet.cell(row=12, column=1).font = Font(bold=True)
                stats_sheet.cell(row=14, column=1).font = Font(bold=True)
                stats_sheet.cell(row=15, column=1).font = Font(bold=True)
                stats_sheet.cell(row=17, column=1).font = Font(bold=True)
                stats_sheet.cell(row=18, column=1).font = Font(bold=True)

                # Count all compression periods and pauses
                compression_period_list = []
                not_available_list = []
                pause_list = []
                for h in range(2, compression_worksheet.max_row + 1):
                    if compression_worksheet.cell(row=h, column=10).value != "FALSE" and \
                            compression_worksheet.cell(row=h, column=14).value != "Artifact":
                        if compression_worksheet.cell(row=h, column=11).value == "N/A":
                            not_available_list.append(compression_worksheet.cell(row=h, column=11).value)
                        else:
                            compression_period_list.append(compression_worksheet.cell(row=h, column=11).value)
                        if compression_worksheet.cell(row=h, column=13).value == "TRUE":
                            pause_list.append(compression_worksheet.cell(row=h, column=11).value)
                compression_period_list.sort()
                pause_list.sort()

                # Calculate needed statistics
                compression_count = len(compression_period_list)
                raw_compression_count = compression_count + len(not_available_list)
                pause_count = len(pause_list)

                cp_mean = get_mean_value(compression_period_list)
                pause_mean = get_mean_value(pause_list)

                cp_minimum = get_minimum_value(compression_period_list)
                pause_minimum = get_minimum_value(pause_list)

                cp_maximum = get_maximum_value(compression_period_list)
                pause_maximum = get_maximum_value(pause_list)

                try:
                    cp_std_dev = statistics.stdev(compression_period_list, cp_mean)
                except statistics.StatisticsError:
                    cp_std_dev = 0
                try:
                    pause_std_dev = statistics.stdev(pause_list, pause_mean)
                except statistics.StatisticsError:
                    pause_std_dev = 0

                cp_variance = math.pow(cp_std_dev, 2)
                pause_variance = math.pow(pause_std_dev, 2)

                try:
                    cp_median = get_median_value(compression_period_list)
                except IndexError:
                    cp_median = 0
                try:
                    pause_median = get_median_value(pause_list)
                except IndexError:
                    pause_median = 0

                try:
                    cp_int_range = get_interquartile_range_values(compression_period_list, cp_median)
                except IndexError:
                    cp_int_range = 0
                try:
                    pause_int_range = get_interquartile_range_values(pause_list, pause_median)
                except IndexError:
                    pause_int_range = 0

                try:
                    cp_std_error = cp_std_dev / math.sqrt(compression_count)
                except ZeroDivisionError:
                    cp_std_error = 0
                try:
                    pause_std_error = pause_std_dev / math.sqrt(pause_count)
                except ZeroDivisionError:
                    pause_std_error = 0

                # Paste statistics into sheet
                stats_sheet.cell(row=1, column=2).value = case

                stats_sheet.cell(row=4, column=2).value = cp_mean
                stats_sheet.cell(row=4, column=3).value = cp_minimum
                stats_sheet.cell(row=4, column=4).value = cp_maximum
                stats_sheet.cell(row=4, column=5).value = cp_std_dev
                stats_sheet.cell(row=4, column=6).value = cp_variance
                stats_sheet.cell(row=4, column=7).value = cp_median
                stats_sheet.cell(row=4, column=8).value = cp_int_range
                stats_sheet.cell(row=4, column=9).value = cp_std_error

                stats_sheet.cell(row=5, column=2).value = cp_mean / 1000
                stats_sheet.cell(row=5, column=3).value = cp_minimum / 1000
                stats_sheet.cell(row=5, column=4).value = cp_maximum / 1000
                stats_sheet.cell(row=5, column=5).value = cp_std_dev / 1000
                stats_sheet.cell(row=5, column=6).value = math.pow(cp_std_dev / 1000, 2)
                stats_sheet.cell(row=5, column=7).value = cp_median / 1000
                stats_sheet.cell(row=5, column=8).value = cp_int_range / 1000
                try:
                    stats_sheet.cell(row=5, column=9).value = (cp_std_dev / 1000) / math.sqrt(compression_count)
                except ZeroDivisionError:
                    stats_sheet.cell(row=5, column=9).value = 0

                stats_sheet.cell(row=8, column=2).value = pause_mean
                stats_sheet.cell(row=8, column=3).value = pause_minimum
                stats_sheet.cell(row=8, column=4).value = pause_maximum
                stats_sheet.cell(row=8, column=5).value = pause_std_dev
                stats_sheet.cell(row=8, column=6).value = pause_variance
                stats_sheet.cell(row=8, column=7).value = pause_median
                stats_sheet.cell(row=8, column=8).value = pause_int_range
                stats_sheet.cell(row=8, column=9).value = pause_std_error

                stats_sheet.cell(row=9, column=2).value = pause_mean / 1000
                stats_sheet.cell(row=9, column=3).value = pause_minimum / 1000
                stats_sheet.cell(row=9, column=4).value = pause_maximum / 1000
                stats_sheet.cell(row=9, column=5).value = pause_std_dev / 1000
                stats_sheet.cell(row=9, column=6).value = math.pow(pause_std_dev / 1000, 2)
                stats_sheet.cell(row=9, column=7).value = pause_median / 1000
                stats_sheet.cell(row=9, column=8).value = pause_int_range / 1000
                try:
                    stats_sheet.cell(row=9, column=9).value = (pause_std_dev / 1000) / math.sqrt(pause_count)
                except ZeroDivisionError:
                    stats_sheet.cell(row=9, column=9).value = 0

                stats_sheet.cell(row=11, column=2).value = raw_compression_count

                stats_sheet.cell(row=14, column=2).value = compression_count

                stats_sheet.cell(row=17, column=2).value = pause_count

                # Save New Compression Data File
                save_path = new_compression_file_path + "\\" + case + ".xlsx"
                new_compression_workbook.save(filename=save_path)
                new_compression_workbook.close()

                # Paste case's statistics into master file
                master_worksheet.cell(row=j + 2, column=1).value = case
                master_worksheet.cell(row=j + 2, column=2).value = raw_compression_count
                master_worksheet.cell(row=j + 2, column=3).value = compression_count
                master_worksheet.cell(row=j + 2, column=4).value = pause_count
                master_worksheet.cell(row=j + 2, column=5).value = cp_mean
                master_worksheet.cell(row=j + 2, column=6).value = cp_minimum
                master_worksheet.cell(row=j + 2, column=7).value = cp_maximum
                master_worksheet.cell(row=j + 2, column=8).value = cp_std_dev
                master_worksheet.cell(row=j + 2, column=9).value = cp_variance
                master_worksheet.cell(row=j + 2, column=10).value = cp_median
                master_worksheet.cell(row=j + 2, column=11).value = cp_int_range
                master_worksheet.cell(row=j + 2, column=12).value = cp_std_error
                master_worksheet.cell(row=j + 2, column=13).value = pause_mean
                master_worksheet.cell(row=j + 2, column=14).value = pause_minimum
                master_worksheet.cell(row=j + 2, column=15).value = pause_maximum
                master_worksheet.cell(row=j + 2, column=16).value = pause_std_dev
                master_worksheet.cell(row=j + 2, column=17).value = pause_variance
                master_worksheet.cell(row=j + 2, column=18).value = pause_median
                master_worksheet.cell(row=j + 2, column=19).value = pause_int_range
                master_worksheet.cell(row=j + 2, column=20).value = pause_std_error

                seconds_sheet.cell(row=j + 2, column=1).value = case
                seconds_sheet.cell(row=j + 2, column=2).value = raw_compression_count
                seconds_sheet.cell(row=j + 2, column=3).value = compression_count
                seconds_sheet.cell(row=j + 2, column=4).value = pause_count
                seconds_sheet.cell(row=j + 2, column=5).value = cp_mean / 1000
                seconds_sheet.cell(row=j + 2, column=6).value = cp_minimum / 1000
                seconds_sheet.cell(row=j + 2, column=7).value = cp_maximum / 1000
                seconds_sheet.cell(row=j + 2, column=8).value = cp_std_dev / 1000
                seconds_sheet.cell(row=j + 2, column=9).value = math.pow(cp_std_dev / 1000, 2)
                seconds_sheet.cell(row=j + 2, column=10).value = cp_median / 1000
                seconds_sheet.cell(row=j + 2, column=11).value = cp_int_range / 1000
                try:
                    seconds_sheet.cell(row=j + 2, column=12).value = (cp_std_dev / 1000) / math.sqrt(compression_count)
                except ZeroDivisionError:
                    seconds_sheet.cell(row=j + 2, column=12).value = 0
                seconds_sheet.cell(row=j + 2, column=13).value = pause_mean / 1000
                seconds_sheet.cell(row=j + 2, column=14).value = pause_minimum / 1000
                seconds_sheet.cell(row=j + 2, column=15).value = pause_maximum / 1000
                seconds_sheet.cell(row=j + 2, column=16).value = pause_std_dev / 1000
                seconds_sheet.cell(row=j + 2, column=17).value = math.pow(pause_std_dev / 1000, 2)
                seconds_sheet.cell(row=j + 2, column=18).value = pause_median / 1000
                seconds_sheet.cell(row=j + 2, column=19).value = pause_int_range / 1000
                try:
                    seconds_sheet.cell(row=j + 2, column=20).value = (pause_std_dev / 1000) / math.sqrt(pause_count)
                except ZeroDivisionError:
                    seconds_sheet.cell(row=j + 2, column=20).value = 0

        # Add missing files to the appropriate lists
        else:
            in_excel_not_cpr.append(case)

    # Add missing files to the appropriate lists
    for case_number in cpr_period_list:
        if case_number not in final_compression_file_list:
            in_cpr_not_excel.append(case_number)

    # Add a sheet into the master workbook to show all files which couldn't be processed
    master_workbook.create_sheet("Missing Case Files")
    missing_files_sheet = master_workbook.worksheets[2]
    missing_files_sheet.cell(row=1, column=1).value = "Cases Missing CPR_Period"
    missing_files_sheet.cell(row=1, column=2).value = "Cases Missing Compression Data"
    for p in range(1, 3):
        missing_files_sheet.cell(row=1, column=p).font = Font(bold=True)
        missing_files_sheet.column_dimensions[get_column_letter(p)].width = 38
    for b in range(0, len(in_excel_not_cpr)):
        missing_files_sheet.cell(row=b+2, column=1).value = in_excel_not_cpr[b]
    for c in range(0, len(in_cpr_not_excel)):
        missing_files_sheet.cell(row=c+2, column=2).value = in_cpr_not_excel[c]

    # Save Master Data File
    save_path = path + "\\" + "Compression_Pause_Master_Data_File.xlsx"
    master_workbook.save(filename=save_path)
    master_workbook.close()
    print("\nCompression Pause Data Workbook saved to " + save_path + ".")


# Remove un-needed characters from the end of the file name
def remove_extra_characters(file_name, file_type):
    reversed_file_name = ""
    # Reverse string to find last instance of '_'
    for k in range(len(file_name), 0, -1):
        reversed_file_name += file_name[k - 1]
    underscore_marker = reversed_file_name.find("_")
    # If no '_', just eliminate file extension
    if underscore_marker == -1:
        # Remove only 4 characters if it is a text file, and 5 if it is an Excel file
        if file_type == "Text":
            return file_name.replace(file_name[len(file_name) - 4: len(file_name)], "")
        else:
            return file_name.replace(file_name[len(file_name) - 5: len(file_name)], "")
    # If '_' has a character to its left, it doesn't have extra digits
    else:
        try:
            # If there are extra digits, remove them along with the file extension
            char_or_int = int(file_name[len(file_name) - 1 - underscore_marker - 1])
            return file_name.replace(file_name[len(file_name) - 1 - underscore_marker: len(file_name)], "")
        except ValueError:
            if file_type == "Text":
                return file_name.replace(file_name[len(file_name) - 4: len(file_name)], "")
            else:
                return file_name.replace(file_name[len(file_name) - 5: len(file_name)], "")


# Calculate Mean Value
def get_mean_value(data):
    try:
        return float(sum(data) / len(data))
    except ZeroDivisionError:
        return 0


# Calculate Minimum Value
def get_minimum_value(data):
    try:
        return data[0]
    except IndexError:
        return 0


# Calculate Maximum Value
def get_maximum_value(data):
    try:
        return data[len(data) - 1]
    except IndexError:
        return 0


# Calculate Median Values
def get_median_value(data):
    # Process is different depending on if the length of the data is even or odd
    if len(data) % 2 == 0:
        upper_value = data[math.ceil(len(data)/2)]
        lower_value = data[math.ceil(len(data)/2) - 1]
        return float((lower_value + upper_value) / 2)
    else:
        return data[math.ceil(len(data)/2) - 1]


# Calculate Interquartile Range Values
def get_interquartile_range_values(data, median):
    # Split data between lower and upper half
    lower_half = []
    upper_half = []
    for value in data:
        if value < median:
            lower_half.append(value)
        elif value > median:
            upper_half.append(value)
        else:
            pass  # If value = median, don't include it in either list
    q1_range = get_median_value(lower_half)
    q3_range = get_median_value(upper_half)
    return q3_range - q1_range


# Re-work "Remove Extra Characters" Function
def remove_extra_characters_2(file_name):
    file_name = file_name[0: len(file_name) - 5]
    file_name_tail = file_name[len(file_name) - 3: len(file_name)]
    if file_name_tail == "_01" or file_name_tail == "_02":
        return file_name[0: len(file_name) - 3]
    else:
        return file_name


start_time = time.time()
find_compression_pauses(r"C:\Users\mnarcisi\Documents\Mike\Scientific Affairs")
# Print time taken to run script
elapsed_time = time.time() - start_time
print("\nTotal time to run script: " + str(round(elapsed_time, 3)) + " seconds.")
